import os
import shutil
import tempfile
from typing import Optional
from fastapi import APIRouter, File, Form, UploadFile, HTTPException, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

from db.database import get_db
from db.models import Projeto, Ambiente
from api.schemas import ProjetoOut, ProjetoListItem, AmbienteUpdate
from parser.dxf_parser import parse_dxf
from parser.pdf_extractor import extract_pdf_measurements
from parser.cross_reference import merge, merge_pdf_only

router = APIRouter()

_FLAG_LABEL = {"confirmed": "✓", "estimated": "⚠️", "missing": "✗"}
_MAX_MB = int(os.getenv("MAX_UPLOAD_MB", "50"))


def _save_temp(upload: UploadFile, suffix: str) -> str:
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    shutil.copyfileobj(upload.file, tmp)
    tmp.close()
    return tmp.name


def _check_size(upload: UploadFile):
    upload.file.seek(0, 2)
    size_mb = upload.file.tell() / (1024 * 1024)
    upload.file.seek(0)
    if size_mb > _MAX_MB:
        raise HTTPException(400, f"Arquivo maior que {_MAX_MB}MB")


def _salvar_ambientes(db: Session, projeto_id: str, ambientes: list[dict]):
    for i, a in enumerate(ambientes):
        amb = Ambiente(
            projeto_id=projeto_id,
            nome=a.get("nome"),
            nome_flag=a.get("nome_flag", "missing"),
            nome_fonte="auto",
            area=a.get("area"),
            area_flag=a.get("area_flag", "missing"),
            perimetro=a.get("perimetro"),
            perimetro_flag=a.get("perimetro_flag", "missing"),
            pe_direito=a.get("pe_direito"),
            pe_direito_flag=a.get("pe_direito_flag", "missing"),
            comprimento=a.get("comprimento"),
            largura=a.get("largura"),
            camada=a.get("camada"),
            fonte=a.get("fonte", "dxf"),
            ordem=i,
        )
        db.add(amb)
    db.commit()


# ── Upload ──────────────────────────────────────────────────────────────────

@router.post("/upload", response_model=ProjetoOut)
async def upload(
    dxf_file: Optional[UploadFile] = File(None),
    pdf_file: Optional[UploadFile] = File(None),
    nome: str = Form("Projeto sem nome"),
    unidade: str = Form("mm"),
    db: Session = Depends(get_db),
):
    if not dxf_file and not pdf_file:
        raise HTTPException(400, "Envie pelo menos um arquivo DXF ou PDF")

    dxf_path = pdf_path = None
    ambientes = []
    fidelidade = None
    confianca = None
    modo = "dxf"

    try:
        if dxf_file:
            _check_size(dxf_file)
            dxf_path = _save_temp(dxf_file, ".dxf")

        if pdf_file:
            _check_size(pdf_file)
            pdf_path = _save_temp(pdf_file, ".pdf")

        # --- Processamento ---
        if dxf_path and pdf_path:
            modo = "dxf+pdf"
            dxf_ambientes = parse_dxf(dxf_path, unidade)
            pdf_ambientes, confianca, _ = extract_pdf_measurements(pdf_path, unidade)
            ambientes, fidelidade = merge(dxf_ambientes, pdf_ambientes)

        elif dxf_path:
            modo = "dxf"
            ambientes = parse_dxf(dxf_path, unidade)

        else:
            modo = "pdf"
            pdf_ambientes, confianca, _ = extract_pdf_measurements(pdf_path, unidade)
            ambientes = merge_pdf_only(pdf_ambientes)

        # --- Salva no banco ---
        projeto = Projeto(
            nome=nome,
            arquivo_dxf=dxf_file.filename if dxf_file else None,
            arquivo_pdf=pdf_file.filename if pdf_file else None,
            unidade=unidade,
            modo=modo,
            fidelidade_score=fidelidade,
            confianca_score=confianca,
        )
        db.add(projeto)
        db.flush()
        _salvar_ambientes(db, projeto.id, ambientes)
        db.refresh(projeto)
        return projeto

    finally:
        for p in [dxf_path, pdf_path]:
            if p and os.path.exists(p):
                os.unlink(p)


# ── Leitura ──────────────────────────────────────────────────────────────────

@router.get("/projetos", response_model=list[ProjetoListItem])
def listar_projetos(db: Session = Depends(get_db)):
    projetos = db.query(Projeto).order_by(Projeto.criado_em.desc()).all()
    return [
        ProjetoListItem(
            id=p.id,
            nome=p.nome,
            modo=p.modo,
            criado_em=p.criado_em.strftime("%d/%m/%Y %H:%M"),
        )
        for p in projetos
    ]


@router.get("/projeto/{projeto_id}", response_model=ProjetoOut)
def get_projeto(projeto_id: str, db: Session = Depends(get_db)):
    p = db.query(Projeto).filter(Projeto.id == projeto_id).first()
    if not p:
        raise HTTPException(404, "Projeto não encontrado")
    return p


# ── Edição manual ─────────────────────────────────────────────────────────────

@router.patch("/ambiente/{ambiente_id}")
def atualizar_ambiente(
    ambiente_id: str,
    body: AmbienteUpdate,
    db: Session = Depends(get_db),
):
    amb = db.query(Ambiente).filter(Ambiente.id == ambiente_id).first()
    if not amb:
        raise HTTPException(404, "Ambiente não encontrado")

    if body.nome is not None:
        amb.nome = body.nome
        amb.nome_flag = "confirmed"
        amb.nome_fonte = "manual"
        amb.fonte = "manual"
    if body.area is not None:
        amb.area = body.area
        amb.area_flag = "confirmed"
    if body.perimetro is not None:
        amb.perimetro = body.perimetro
        amb.perimetro_flag = "confirmed"
    if body.pe_direito is not None:
        amb.pe_direito = body.pe_direito
        amb.pe_direito_flag = "confirmed"

    db.commit()
    db.refresh(amb)
    return {"ok": True}


# ── Export Excel ──────────────────────────────────────────────────────────────

@router.get("/projeto/{projeto_id}/export")
def exportar_xlsx(projeto_id: str, db: Session = Depends(get_db)):
    p = db.query(Projeto).filter(Projeto.id == projeto_id).first()
    if not p:
        raise HTTPException(404, "Projeto não encontrado")

    wb = Workbook()
    ws = wb.active
    ws.title = "Ambientes"

    # Cabeçalho
    headers = ["Ambiente", "Camada", "Área (m²)", "C (m)", "L (m)", "Perímetro (m)", "Pé Direito (m)", "Fonte"]
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Dados
    FLAG_SUFFIX = {"confirmed": "", "estimated": " ⚠️", "missing": " ✗"}

    for row, amb in enumerate(sorted(p.ambientes, key=lambda x: x.ordem), 2):
        nome_val = (amb.nome or "— sem nome —") + FLAG_SUFFIX.get(amb.nome_flag, "")
        area_val = f"{amb.area:.2f}{FLAG_SUFFIX.get(amb.area_flag, '')}" if amb.area else "✗"
        c_val = f"{amb.comprimento:.2f}" if amb.comprimento else "—"
        l_val = f"{amb.largura:.2f}" if amb.largura else "—"
        perim_val = f"{amb.perimetro:.2f}{FLAG_SUFFIX.get(amb.perimetro_flag, '')}" if amb.perimetro else "✗"
        pd_val = f"{amb.pe_direito:.2f}{FLAG_SUFFIX.get(amb.pe_direito_flag, '')}" if amb.pe_direito else "✗"

        ws.cell(row=row, column=1, value=nome_val)
        ws.cell(row=row, column=2, value=amb.camada or "—")
        ws.cell(row=row, column=3, value=area_val)
        ws.cell(row=row, column=4, value=c_val)
        ws.cell(row=row, column=5, value=l_val)
        ws.cell(row=row, column=6, value=perim_val)
        ws.cell(row=row, column=7, value=pd_val)
        ws.cell(row=row, column=8, value=amb.fonte)

        # Linha alternada
        if row % 2 == 0:
            fill = PatternFill("solid", fgColor="F0F4F8")
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = fill

    # Ajusta largura das colunas
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Linha de resumo
    total_row = ws.max_row + 2
    ws.cell(row=total_row, column=1, value="TOTAL GERAL").font = Font(bold=True)
    areas = [a.area for a in p.ambientes if a.area]
    ws.cell(row=total_row, column=3, value=f"{sum(areas):.2f} m²").font = Font(bold=True)

    # Metadados
    meta_row = total_row + 2
    ws.cell(row=meta_row, column=1, value=f"Projeto: {p.nome}")
    ws.cell(row=meta_row + 1, column=1, value=f"Unidade original: {p.unidade}")
    if p.fidelidade_score:
        ws.cell(row=meta_row + 2, column=1, value=f"Fidelidade DXF×PDF: {p.fidelidade_score}%")
    if p.confianca_score:
        ws.cell(row=meta_row + 3, column=1, value=f"Confiança PDF: {p.confianca_score}%")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"cotacad_{p.nome.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
